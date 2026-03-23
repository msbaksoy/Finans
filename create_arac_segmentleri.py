#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Araç segmentleri Excel dosyası oluşturucu - Hedef Filo blog verisine göre"""

import os

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    import subprocess
    subprocess.check_call(["pip", "install", "openpyxl"])
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

# Segment açıklamaları (Hedef Filo blog)
SEGMENT_ACIKLAMALARI = {
    "A": "En küçük araçlar (max 3,7 m). Şehir içi, park kolaylığı, uygun fiyat.",
    "B": "3,7–4 m arası küçük araçlar. Uygun fiyat, A'ya göre daha ağır ve güçlü.",
    "C": "Kompakt / alt-orta sınıf. Küçük aile otomobili. Türkiye'de en çok satan segment.",
    "D": "Üst-orta sınıf / geniş aile aracı. Geniş iç mekan, yüksek silindir hacmi.",
    "E": "Üst sınıf, 5 m+ lüks araçlar. 2.0 CC+ motor, ağır ve kaliteli malzeme.",
    "F": "Lüks otomobil sınıfı. Geniş hacim, yüksek güç, üstün işçilik ve donanım.",
    "G": "Spor araçlar. Üstün performans, cabrio/coupe/roadster karoser.",
    "J": "SUV ve CUV. 4x4, off-road, şehir içi uyumlu.",
    "M": "MPV – çok amaçlı araçlar. Birden fazla segment özelliği.",
    "S": "İki kapılı spor. Yüksek beygir, üstün manevra ve yol tutuş.",
}

# Segment | Marka | Model (Hedef Filo tablosundan)
ARACLAR = [
    ("A", "Chevrolet", "Spark"),
    ("A", "Citroen", "C1"),
    ("A", "Fiat", "500"),
    ("A", "Fiat", "Panda"),
    ("A", "Ford", "Ka"),
    ("A", "Hyundai", "i10"),
    ("A", "Kia", "Picanto"),
    ("A", "Opel", "Adam"),
    ("A", "Peugeot", "107"),
    ("A", "Skoda", "Citigo"),
    ("A", "Subaru", "R1"),
    ("A", "Suzuki", "Alto"),
    ("A", "Volkswagen", "Up"),
    ("B", "Audi", "A1"),
    ("B", "Chevrolet", "Kalos"),
    ("B", "Citroen", "C Elysee"),
    ("B", "Citroen", "C3"),
    ("B", "Dacia", "Logan"),
    ("B", "Dacia", "Sandero"),
    ("B", "Fiat", "Albea"),
    ("B", "Fiat", "Punto"),
    ("B", "Fiat", "Palio"),
    ("B", "Ford", "Fiesta"),
    ("B", "Honda", "Jazz"),
    ("B", "Honda", "City"),
    ("B", "Hyundai", "i20"),
    ("B", "Hyundai", "Getz"),
    ("B", "Hyundai", "Accent Blue"),
    ("B", "Kia", "Rio"),
    ("B", "Mazda", "2"),
    ("B", "Mitsubishi", "Attrage"),
    ("B", "Nissan", "Juke"),
    ("B", "Nissan", "Micra"),
    ("B", "Nissan", "Note"),
    ("B", "Opel", "Corsa"),
    ("B", "Opel", "Meriva"),
    ("B", "Opel", "Mokka"),
    ("B", "Peugeot", "206"),
    ("B", "Peugeot", "208"),
    ("B", "Peugeot", "301"),
    ("B", "Peugeot", "2008"),
    ("B", "Renault", "Clio"),
    ("B", "Renault", "Symbol"),
    ("B", "Seat", "İbiza"),
    ("B", "Seat", "Toledo"),
    ("B", "Seat", "Cordoba"),
    ("B", "Seat", "Exeo"),
    ("B", "Skoda", "Rapid"),
    ("B", "Skoda", "Fabia"),
    ("B", "Subaru", "Justy"),
    ("B", "Suzuki", "Swift"),
    ("B", "Toyota", "Yaris"),
    ("B", "Volkswagen", "Polo"),
    ("C", "Alfa Romeo", "Giulietta"),
    ("C", "Audi", "A3"),
    ("C", "Audi", "Q3"),
    ("C", "Audi", "Q5"),
    ("C", "BMW", "1 Serisi"),
    ("C", "BMW", "2 Serisi"),
    ("C", "Chevrolet", "Cruze"),
    ("C", "Citroen", "C4"),
    ("C", "Citroen", "C4 Picasso"),
    ("C", "Dacia", "Lodgy"),
    ("C", "Dacia", "Duster"),
    ("C", "Fiat", "Bravo"),
    ("C", "Fiat", "Linea"),
    ("C", "Fiat", "Egea"),
    ("C", "Ford", "Focus"),
    ("C", "Ford", "C-Max"),
    ("C", "Honda", "Civic"),
    ("C", "Hyundai", "i30"),
    ("C", "Hyundai", "Elantra"),
    ("C", "Hyundai", "Tucson"),
    ("C", "Kia", "Ceed"),
    ("C", "Kia", "Cerato"),
    ("C", "Kia", "Sportage"),
    ("C", "Mazda", "3"),
    ("C", "Mercedes", "A Serisi"),
    ("C", "Mercedes", "CLA"),
    ("C", "Mercedes", "GLA"),
    ("C", "Mitsubishi", "Lancer"),
    ("C", "Nissan", "Qashqai"),
    ("C", "Opel", "Astra"),
    ("C", "Peugeot", "308"),
    ("C", "Peugeot", "3008"),
    ("C", "Renault", "Fluence"),
    ("C", "Renault", "Megane"),
    ("C", "Renault", "Scenic"),
    ("C", "Renault", "Kadjar"),
    ("C", "Seat", "Leon"),
    ("C", "Skoda", "Octavia"),
    ("C", "Subaru", "Impreza"),
    ("C", "Toyota", "Auris"),
    ("C", "Toyota", "Corolla"),
    ("C", "Toyota", "Verso"),
    ("C", "Toyota", "Corona"),
    ("C", "Volkswagen", "Golf"),
    ("C", "Volkswagen", "Jetta"),
    ("C", "Volkswagen", "Tiguan"),
    ("C", "Volvo", "V40"),
    ("D", "Alfa Romeo", "156"),
    ("D", "Alfa Romeo", "159"),
    ("D", "Alfa Romeo", "Giulia"),
    ("D", "Audi", "A4"),
    ("D", "Audi", "A5"),
    ("D", "BMW", "3 Serisi"),
    ("D", "BMW", "4 Serisi"),
    ("D", "BMW", "X3"),
    ("D", "Citroen", "Grand C4 Picasso"),
    ("D", "Citroen", "C5"),
    ("D", "Ford", "Mondeo"),
    ("D", "Ford", "S-Max"),
    ("D", "Ford", "Galaxy"),
    ("D", "Honda", "CRV"),
    ("D", "Honda", "Accord"),
    ("D", "Hyundai", "Sonata"),
    ("D", "Hyundai", "Santa Fe"),
    ("D", "Kia", "Optima"),
    ("D", "Kia", "Sorento"),
    ("D", "Mazda", "6"),
    ("D", "Mercedes", "C Serisi"),
    ("D", "Nissan", "X-Trail"),
    ("D", "Nissan", "Primera"),
    ("D", "Opel", "Insignia"),
    ("D", "Opel", "Zafira"),
    ("D", "Opel", "Vectra"),
    ("D", "Peugeot", "508"),
    ("D", "Renault", "Laguna"),
    ("D", "Renault", "Talisman"),
    ("D", "Skoda", "Superb"),
    ("D", "Toyota", "Avensis"),
    ("D", "Volkswagen", "Passat"),
    ("D", "Volvo", "S60"),
    ("E", "Audi", "A6"),
    ("E", "Audi", "A7"),
    ("E", "Audi", "Q7"),
    ("E", "BMW", "5 Serisi"),
    ("E", "BMW", "X5"),
    ("E", "BMW", "X6"),
    ("E", "Ford", "Ranger"),
    ("E", "Jaguar", "XF"),
    ("E", "Mercedes", "E Serisi"),
    ("E", "Nissan", "Navara"),
    ("E", "Saab", "9-5"),
    ("E", "Volkswagen", "Amarok"),
    ("E", "Volvo", "S80"),
    ("F", "Audi", "A8"),
    ("F", "Bentley", "Bentayga"),
    ("F", "Bentley", "Mulsanne"),
    ("F", "BMW", "7 Serisi"),
    ("F", "Jaguar", "XJ"),
    ("F", "Jeep", "Grand Cherokee"),
    ("F", "Land Rover", "Range Rover"),
    ("F", "Maserati", "Levante"),
    ("F", "Maserati", "Quattroporte"),
    ("F", "Mercedes", "G-Serisi"),
    ("F", "Mercedes", "GLE Serisi"),
    ("F", "Mercedes", "S Serisi"),
    ("F", "Mitsubishi", "Pajero"),
    ("F", "Porsche", "Cayenne"),
    ("F", "Rolls-Royce", "Cullinan"),
    ("F", "Rolls-Royce", "Ghost"),
    ("F", "Volkswagen", "Phaeton"),
    ("F", "Volvo", "XC90"),
    ("G", "Alfa Romeo", "4C Spider"),
    ("G", "Alfa Romeo", "Spider"),
    ("G", "Aston Martin", "Lagonda"),
    ("G", "Audi", "TT"),
    ("G", "BMW", "Z4"),
    ("G", "Fiat", "124 Spider"),
    ("G", "Honda", "S2000"),
    ("G", "Jaguar", "F Type"),
    ("G", "Lotus", "Elise"),
    ("G", "Mazda", "MX-5"),
    ("G", "Mercedes", "SLC Serisi"),
    ("G", "Porsche", "911"),
    ("G", "Porsche", "718 Boxster"),
    ("G", "Tesla", "Roadster"),
    ("J", "Alfa Romeo", "Stelvio"),
    ("J", "Audi", "Q3"),
    ("J", "Audi", "Q5"),
    ("J", "Audi", "Q7"),
    ("J", "BMW", "X1"),
    ("J", "BMW", "X3"),
    ("J", "BMW", "X5"),
    ("J", "BMW", "X6"),
    ("J", "Fiat", "500X"),
    ("J", "Fiat", "Freemont"),
    ("J", "Ford", "Edge"),
    ("J", "Hyundai", "Santa Fe"),
    ("J", "Hyundai", "Santa Fe Sport"),
    ("J", "Infiniti", "QX50"),
    ("J", "Infiniti", "QX60"),
    ("J", "Jaguar", "F-Pace"),
    ("J", "Jeep", "Cherokee"),
    ("J", "Kia", "Sorento"),
    ("J", "Land Rover", "Discovery Sport"),
    ("J", "Land Rover", "Range Rover Velar"),
    ("J", "Mazda", "CX-9"),
    ("J", "Mercedes", "GLC Serisi"),
    ("J", "Mercedes", "GLC Serisi Coupe"),
    ("J", "Nissan", "X-Trail"),
    ("J", "Peugeot", "5008"),
    ("J", "Porsche", "Macan"),
    ("J", "Renault", "Koleos"),
    ("J", "Seat", "Ateca XL"),
    ("J", "Skoda", "Kodiaq"),
    ("J", "Volkswagen", "Tiguan Allspace"),
    ("J", "Volvo", "XC60"),
    ("M", "BMW", "2 Serisi Active Tourer"),
    ("M", "BMW", "2 Serisi Gran Tourer"),
    ("M", "Citroen", "C4 Picasso"),
    ("M", "Citroen", "Grand C4 Picasso"),
    ("M", "Fiat", "500L"),
    ("M", "Ford", "B-Max"),
    ("M", "Ford", "C-Max"),
    ("M", "Ford", "Galaxy"),
    ("M", "Ford", "Grand C-Max"),
    ("M", "Ford", "S-Max"),
    ("M", "Hyundai", "iX20"),
    ("M", "Kia", "Carens"),
    ("M", "Kia", "Sedona"),
    ("M", "Kia", "Venga"),
    ("M", "Lancia", "Musa"),
    ("M", "Mazda", "5"),
    ("M", "Mercedes", "B Serisi"),
    ("M", "Mercedes", "V Serisi"),
    ("M", "Mini", "Clubman"),
    ("M", "Nissan", "Note"),
    ("M", "Opel", "Meriva"),
    ("M", "Opel", "Zafira"),
    ("M", "Opel", "Zafira Tourer"),
    ("M", "Renault", "Espace"),
    ("M", "Renault", "Scenic"),
    ("M", "Renault", "Grand Scenic"),
    ("M", "Renault", "Modus"),
    ("M", "Seat", "Alhambra"),
    ("M", "Seat", "Altea"),
    ("M", "Seat", "Altea XL"),
    ("M", "Skoda", "Roomster"),
    ("M", "Suzuki", "Ertiga"),
    ("M", "Toyota", "Verso"),
    ("M", "Toyota", "Verso-S"),
    ("M", "Volkswagen", "Caravelle"),
    ("M", "Volkswagen", "Golf Sportsvan"),
    ("M", "Volkswagen", "Sharan"),
    ("M", "Volkswagen", "Touran"),
    ("S", "Alfa Romeo", "4C"),
    ("S", "Alfa Romeo", "8C Competizione"),
    ("S", "Alfa Romeo", "8C Spider"),
    ("S", "Alfa Romeo", "Brera"),
    ("S", "Aston Martin", "V8 Vantage S"),
    ("S", "Audi", "R8 V10 Plus"),
    ("S", "Chevrolet", "Corvette ZR"),
    ("S", "Ferrari", "488 GTB"),
    ("S", "Jaguar", "F Type Coupe"),
    ("S", "Lamborghini", "Huracan LP 610-4"),
    ("S", "Lexus", "LFA"),
    ("S", "Mazda", "RX-7"),
    ("S", "McLaren", "675LT"),
    ("S", "Mercedes", "SLC Serisi"),
    ("S", "Mercedes", "SLS AMG GT"),
    ("S", "Nissan", "370Z"),
    ("S", "Nissan", "GTR Nismo"),
    ("S", "Porsche", "718 Cayman"),
    ("S", "Porsche", "911 GT2 RS"),
    ("S", "Subaru", "BRZ"),
    ("S", "Toyota", "GT86"),
]

DESKTOP = os.path.expanduser("~/Desktop")
OUTPUT_PATH = os.path.join(DESKTOP, "araç segmentleri.xlsx")


def main():
    wb = openpyxl.Workbook()
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    segment_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    segment_font = Font(bold=True, size=10)

    # Sayfa 1: Segment Özeti
    ws1 = wb.active
    ws1.title = "Segment Özeti"
    ws1.append(["Segment", "Açıklama"])
    for row in ws1.iter_rows(min_row=1, max_row=1, min_col=1, max_col=2):
        for cell in row:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border
    for seg in "ABCDEFGJMS":
        ws1.append([f"Segment {seg}", SEGMENT_ACIKLAMALARI.get(seg, "")])
    for r in range(2, ws1.max_row + 1):
        for c in range(1, 3):
            ws1.cell(row=r, column=c).border = border
            ws1.cell(row=r, column=c).alignment = Alignment(vertical="top", wrap_text=True)
    ws1.column_dimensions["A"].width = 14
    ws1.column_dimensions["B"].width = 55

    # Sayfa 2: Marka-Model Listesi
    ws2 = wb.create_sheet("Marka ve Modeller", 1)
    ws2.append(["Segment", "Marka", "Model"])
    for cell in ws2[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    for seg, marka, model in ARACLAR:
        ws2.append([seg, marka, model])
    for r in range(2, ws2.max_row + 1):
        for c in range(1, 4):
            cell = ws2.cell(row=r, column=c)
            cell.border = border
            if c == 1:
                cell.fill = segment_fill
                cell.font = segment_font
    ws2.column_dimensions["A"].width = 10
    ws2.column_dimensions["B"].width = 18
    ws2.column_dimensions["C"].width = 28
    ws2.auto_filter.ref = ws2.dimensions

    wb.save(OUTPUT_PATH)
    print(f"Excel dosyası oluşturuldu: {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
